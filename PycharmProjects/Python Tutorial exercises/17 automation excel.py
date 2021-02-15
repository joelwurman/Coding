import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    # cell = sheet['a1']
    # cell = sheet.cell(1, 1)

    # this for loop adds a column on column 4 with a 90% of the value of column 3
    for row in range(2, sheet.max_row + 1):  # from 2 cause we want to ignore the first row
        cell = sheet.cell(row, 3)  # column3
        corrected_price = cell.value * 0.9
        cell_corrected_price = sheet.cell(row, 4)
        cell_corrected_price.value = corrected_price

    # use Reference to select a range of values
    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)





